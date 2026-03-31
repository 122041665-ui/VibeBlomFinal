import csv
import io
from datetime import datetime

from flask import (
    Blueprint, render_template, session, redirect,
    url_for, flash, jsonify, request, make_response,
)
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)

from app.services.api_client import api_get

dashboard_bp = Blueprint("dashboard", __name__, url_prefix="/dashboard")

# ─────────────────────────────────────────
# Constants
# ─────────────────────────────────────────

DEFAULT_REPORT_OPTIONS = {
    "places": [
        {"value": "all",     "label": "Todas las acciones"},
        {"value": "created", "label": "Agregados"},
        {"value": "updated", "label": "Modificaciones"},
        {"value": "deleted", "label": "Eliminados"},
        {"value": "current", "label": "Solo lugares actuales"},
    ],
    "reviews": [
        {"value": "all",     "label": "Todas las acciones"},
        {"value": "created", "label": "Agregados"},
        {"value": "updated", "label": "Modificaciones"},
        {"value": "deleted", "label": "Eliminados"},
    ],
    "users": [
        {"value": "all",     "label": "Todas las acciones"},
        {"value": "created", "label": "Agregados"},
        {"value": "updated", "label": "Modificaciones"},
        {"value": "deleted", "label": "Eliminados"},
    ],
    "approvals": [
        {"value": "all",      "label": "Todas las acciones"},
        {"value": "approved", "label": "Aprobadas"},
        {"value": "rejected", "label": "Rechazadas"},
    ],
}

DEFAULT_REPORT_HEADERS = [
    "id", "nombre", "accion", "realizado_por", "puesto", "fecha", "estado",
]

# ─────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────

def safe_api_json(path: str, default=None, params=None):
    if default is None:
        default = {}
    try:
        response = api_get(path, params=params)
    except Exception:
        return default, None, "connection_error"

    if response.status_code == 401:
        return default, response, "unauthorized"
    if response.status_code == 403:
        return default, response, "forbidden"
    if response.status_code != 200:
        return default, response, "http_error"

    try:
        return response.json(), response, None
    except Exception:
        return default, response, "invalid_json"


def build_dashboard_context(raw_data):
    raw_data = raw_data or {}
    stats          = raw_data.get("stats")          or {}
    recent         = raw_data.get("recent")         or {}
    current_user   = raw_data.get("current_user")   or {}
    recent_activity = raw_data.get("recent_activity") or []

    return {
        "stats": {
            "users":     stats.get("users",     0),
            "places":    stats.get("places",    0),
            "reviews":   stats.get("reviews",   0),
            "favorites": stats.get("favorites", 0),
            "approvals": stats.get("approvals", 0),
        },
        "current_user": {
            "id":    current_user.get("id"),
            "name":  current_user.get("name",  "Usuario"),
            "email": current_user.get("email", "Sin correo"),
            "role":  current_user.get("role",  "staff"),
        },
        "recent": {
            "users":   recent.get("users")   or [],
            "places":  recent.get("places")  or [],
            "reviews": recent.get("reviews") or [],
        },
        "recent_activity":  recent_activity,
        "places_by_type":   raw_data.get("places_by_type")  or [],
        "users_monthly":    raw_data.get("users_monthly")   or [],
        "places_monthly":   raw_data.get("places_monthly")  or [],
        "report_options":   raw_data.get("report_options")  or DEFAULT_REPORT_OPTIONS,
    }


def get_report_filters():
    return {
        "module":     request.args.get("module",     "places"),
        "scope":      request.args.get("scope",      "all"),
        "start_date": request.args.get("start_date", ""),
        "end_date":   request.args.get("end_date",   ""),
    }


def get_module_label(module_value: str) -> str:
    return {
        "places":    "Lugares",
        "reviews":   "Reseñas",
        "users":     "Usuarios",
        "approvals": "Aprobaciones",
    }.get(module_value, "Reporte")


def get_scope_label(module_value: str, scope_value: str, report_options: dict) -> str:
    for item in report_options.get(module_value) or []:
        if item.get("value") == scope_value:
            return item.get("label", scope_value)
    return scope_value


def fetch_report(filters: dict) -> dict:
    """Fetch report data from FastAPI and return normalized dict."""
    raw, _, error = safe_api_json(
        "/admin/dashboard/report-data", default={}, params=filters
    )
    report_options = raw.get("report_options") or DEFAULT_REPORT_OPTIONS
    return {
        "headers":      raw.get("headers") or DEFAULT_REPORT_HEADERS,
        "rows":         raw.get("rows")    or [],
        "total":        raw.get("total",   0),
        "module_label": raw.get("module_label") or get_module_label(filters["module"]),
        "scope_label":  raw.get("scope_label")  or get_scope_label(
            filters["module"], filters["scope"], report_options
        ),
        "error": error,
    }


def safe_str(value) -> str:
    if value is None:
        return "—"
    return str(value).strip() or "—"


# ─────────────────────────────────────────
# PDF generation
# ─────────────────────────────────────────

BRAND_BLUE  = colors.HexColor("#1d4ed8")
BRAND_LIGHT = colors.HexColor("#eff6ff")
BRAND_GREY  = colors.HexColor("#475569")
BRAND_DARK  = colors.HexColor("#0f172a")
ROW_ALT     = colors.HexColor("#f8fafc")


def build_pdf(headers: list, rows: list, module_label: str, scope_label: str,
              start_date: str, end_date: str) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=14 * mm,  bottomMargin=14 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "vb_title",
        parent=styles["Normal"],
        fontSize=18, fontName="Helvetica-Bold",
        textColor=BRAND_BLUE, spaceAfter=2 * mm,
    )
    subtitle_style = ParagraphStyle(
        "vb_sub",
        parent=styles["Normal"],
        fontSize=9, fontName="Helvetica",
        textColor=BRAND_GREY, spaceAfter=4 * mm,
    )
    meta_style = ParagraphStyle(
        "vb_meta",
        parent=styles["Normal"],
        fontSize=8, fontName="Helvetica",
        textColor=BRAND_GREY,
    )

    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    date_range = f"{start_date or '—'} → {end_date or '—'}"

    elements = [
        Paragraph("VibeBloom Admin — Reporte", title_style),
        Paragraph(
            f"{module_label} · {scope_label} · Rango: {date_range} · Generado: {now}",
            subtitle_style,
        ),
        Spacer(1, 4 * mm),
    ]

    # Table data
    col_labels = [h.upper() for h in headers]
    table_data = [col_labels]
    for row in rows:
        table_data.append([safe_str(row.get(h)) for h in headers])

    # Column widths: distribute evenly
    avail_w = landscape(A4)[0] - 28 * mm
    col_w = avail_w / max(len(headers), 1)

    tbl = Table(table_data, colWidths=[col_w] * len(headers), repeatRows=1)

    style_cmds = [
        # Header row
        ("BACKGROUND",   (0, 0), (-1, 0), BRAND_BLUE),
        ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
        ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0), 8),
        ("ALIGN",        (0, 0), (-1, 0), "LEFT"),
        ("TOPPADDING",   (0, 0), (-1, 0), 7),
        ("BOTTOMPADDING",(0, 0), (-1, 0), 7),
        ("LEFTPADDING",  (0, 0), (-1, 0), 8),
        # Data rows
        ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",     (0, 1), (-1, -1), 8),
        ("TEXTCOLOR",    (0, 1), (-1, -1), BRAND_DARK),
        ("ALIGN",        (0, 1), (-1, -1), "LEFT"),
        ("TOPPADDING",   (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING",(0, 1), (-1, -1), 5),
        ("LEFTPADDING",  (0, 1), (-1, -1), 8),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        # Grid
        ("LINEBELOW",    (0, 0), (-1, 0), 0.5, colors.HexColor("#1e40af")),
        ("LINEBELOW",    (0, 1), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ROW_ALT]),
    ]
    tbl.setStyle(TableStyle(style_cmds))
    elements.append(tbl)

    # Footer
    elements += [
        Spacer(1, 6 * mm),
        Paragraph(
            f"VibeBloom Admin · {len(rows)} registro{'s' if len(rows) != 1 else ''} · {now}",
            meta_style,
        ),
    ]

    doc.build(elements)
    return buf.getvalue()


# ─────────────────────────────────────────
# XLS generation
# ─────────────────────────────────────────

def build_xls(headers: list, rows: list, module_label: str, scope_label: str,
              start_date: str, end_date: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = module_label[:31]  # Excel limit

    # Styles
    header_fill = PatternFill("solid", fgColor="1d4ed8")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="left", vertical="center")

    title_font  = Font(name="Calibri", bold=True, color="0f172a", size=14)
    meta_font   = Font(name="Calibri", color="64748b", size=9)
    data_font   = Font(name="Calibri", color="334155", size=10)
    alt_fill    = PatternFill("solid", fgColor="f8fafc")

    thin = Side(style="thin", color="e2e8f0")
    border = Border(bottom=thin)

    now = datetime.now().strftime("%d/%m/%Y %H:%M")

    # Title rows
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=max(len(headers), 1))
    ws.cell(row=1, column=1).value = "VibeBloom Admin — Reporte"
    ws.cell(row=1, column=1).font  = title_font

    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=max(len(headers), 1))
    date_range = f"{start_date or '—'} → {end_date or '—'}"
    ws.cell(row=2, column=1).value = (
        f"{module_label} · {scope_label} · Rango: {date_range} · Generado: {now}"
    )
    ws.cell(row=2, column=1).font  = meta_font

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 15

    # Header row (row 4)
    HEADER_ROW = 4
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx)
        cell.value     = h.upper()
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
    ws.row_dimensions[HEADER_ROW].height = 20

    # Data rows
    for row_idx, row in enumerate(rows, start=HEADER_ROW + 1):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value  = safe_str(row.get(h))
            cell.font   = data_font
            cell.border = border
            if fill:
                cell.fill = fill

    # Auto-width
    for col_idx, h in enumerate(headers, start=1):
        max_len = max(
            len(h),
            max((len(safe_str(r.get(h))) for r in rows), default=0),
        )
        ws.column_dimensions[
            ws.cell(row=HEADER_ROW, column=col_idx).column_letter
        ].width = min(max_len + 4, 40)

    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────
# Routes
# ─────────────────────────────────────────

@dashboard_bp.route("/")
def dashboard():
    if "access_token" not in session:
        return redirect(url_for("auth.login"))

    raw_data, _, error = safe_api_json("/admin/dashboard", default={})

    if error == "connection_error":
        flash("No se pudo conectar con la API central", "error")
        return render_template("dashboard.html", data=build_dashboard_context({}))
    if error == "unauthorized":
        flash("Sesión expirada", "error")
        session.clear()
        return redirect(url_for("auth.login"))
    if error == "forbidden":
        flash("No autorizado", "error")
        return redirect(url_for("auth.login"))
    if error in ("http_error", "invalid_json"):
        flash("No se pudo cargar el dashboard", "error")
        return render_template("dashboard.html", data=build_dashboard_context({}))

    return render_template("dashboard.html", data=build_dashboard_context(raw_data))


@dashboard_bp.route("/report-data")
def report_data():
    if "access_token" not in session:
        return jsonify({"error": "unauthorized"}), 401

    filters = get_report_filters()
    result  = fetch_report(filters)

    if result["error"] == "unauthorized":
        session.clear()
        return jsonify({"error": "unauthorized"}), 401

    return jsonify({
        "headers":      result["headers"],
        "rows":         result["rows"],
        "total":        result["total"],
        "module_label": result["module_label"],
        "scope_label":  result["scope_label"],
    })


@dashboard_bp.route("/export/pdf")
def export_pdf():
    if "access_token" not in session:
        return redirect(url_for("auth.login"))

    filters = get_report_filters()
    result  = fetch_report(filters)

    if result["error"] == "unauthorized":
        flash("Sesión expirada", "error")
        session.clear()
        return redirect(url_for("auth.login"))

    pdf_bytes = build_pdf(
        headers=result["headers"],
        rows=result["rows"],
        module_label=result["module_label"],
        scope_label=result["scope_label"],
        start_date=filters["start_date"],
        end_date=filters["end_date"],
    )

    filename = (
        f"reporte-{result['module_label'].lower()}-"
        f"{datetime.now().strftime('%Y%m%d-%H%M')}.pdf"
    )

    resp = make_response(pdf_bytes)
    resp.headers["Content-Type"] = "application/pdf"
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@dashboard_bp.route("/export/xls")
def export_xls():
    if "access_token" not in session:
        return redirect(url_for("auth.login"))

    filters = get_report_filters()
    result  = fetch_report(filters)

    if result["error"] == "unauthorized":
        flash("Sesión expirada", "error")
        session.clear()
        return redirect(url_for("auth.login"))

    xls_bytes = build_xls(
        headers=result["headers"],
        rows=result["rows"],
        module_label=result["module_label"],
        scope_label=result["scope_label"],
        start_date=filters["start_date"],
        end_date=filters["end_date"],
    )

    filename = (
        f"reporte-{result['module_label'].lower()}-"
        f"{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    )

    resp = make_response(xls_bytes)
    resp.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
